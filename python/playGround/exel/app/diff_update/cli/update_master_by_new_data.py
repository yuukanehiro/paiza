#!/usr/bin/env python3
import os
import shutil
from datetime import datetime
import logging
from openpyxl import load_workbook

# 設定値
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MASTER_FILE = os.path.join(BASE_DIR, "master.xlsx")
DATA_DIR = os.path.join(BASE_DIR, "data")
PROCESSED_DIR = os.path.join(BASE_DIR, "processed")
LOG_DIR = os.path.join(BASE_DIR, "log")
ID_COLUMN = "対象id"  # IDとして使用するカラム名

def setup_logging(log_dir):
    """ログ設定をセットアップ"""
    log_file = os.path.join(log_dir, f"update_master_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return log_file


def get_new_ids(master_data, new_data, id_column):
    """masterにない新しいIDを取得"""
    try:
        master_ids = set()
        for row in master_data:
            if id_column in row and row[id_column] is not None:
                master_ids.add(str(row[id_column]).strip())

        new_ids = set()
        for row in new_data:
            if id_column in row and row[id_column] is not None:
                new_ids.add(str(row[id_column]).strip())

        added_ids = new_ids - master_ids
        logging.info(f"Master ID数: {len(master_ids)}")
        logging.info(f"New Data ID数: {len(new_ids)}")
        logging.info(f"追加対象ID数: {len(added_ids)}")

        return added_ids
    except Exception as e:
        logging.error(f"ID比較エラー: {str(e)}")
        return set()


def move_to_processed(source_file, processed_dir):
    """処理済みファイルをprocessedフォルダに移動"""
    try:
        # processedディレクトリが存在しない場合は作成
        os.makedirs(processed_dir, exist_ok=True)

        # ファイル名に日時を追加
        filename = os.path.basename(source_file)
        name, ext = os.path.splitext(filename)
        processed_filename = f"{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
        processed_path = os.path.join(processed_dir, processed_filename)

        # ファイル移動
        shutil.move(source_file, processed_path)
        logging.info(f"ファイル移動: {source_file} -> {processed_path}")
        return True

    except Exception as e:
        logging.error(f"ファイル移動エラー: {str(e)}")
        return False


def read_excel_data(file_path):
    """Excelファイルを読み込んでデータを返す"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # データを辞書のリストとして読み込み
        data = []
        headers = [cell.value for cell in ws[1]]  # 1行目をヘッダーとして使用

        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):  # 空行をスキップ
                row_data = dict(zip(headers, row))
                data.append(row_data)

        logging.info(f"Excelファイル読み込み成功: {file_path}")
        logging.info(f"データ件数: {len(data)}件")

        return data, headers
    except Exception as e:
        logging.error(f"Excelファイル読み込みエラー: {file_path} - {str(e)}")
        return None, None

def update_master_excel(master_file_path, new_data, new_ids, id_column, headers):
    """masterExcelファイルに新しいデータを追加"""
    try:
        # 新しいIDのデータのみを抽出
        new_records = []
        for row in new_data:
            if id_column in row and str(row[id_column]).strip() in new_ids:
                new_records.append(row)

        if new_records:
            # masterファイルを読み込み
            wb = load_workbook(master_file_path)
            ws = wb.active

            # 新しいレコードを追加
            for record in new_records:
                row_data = [record.get(header, '') for header in headers]
                ws.append(row_data)
                logging.info(f"追加: ID={record[id_column]}")

            # ファイルに保存
            wb.save(master_file_path)
            logging.info(f"Masterファイルを更新: {len(new_records)}件追加")

            return True
        else:
            logging.info("追加するデータがありません")
            return True

    except Exception as e:
        logging.error(f"Masterファイル更新エラー: {str(e)}")
        return False

def process_excel_files():
    """dataフォルダ内のすべてのExcelファイルを処理"""
    # ログセットアップ
    setup_logging(LOG_DIR)

    logging.info("=== Excel データ更新処理開始 ===")

    # masterファイルの読み込み
    master_data, master_headers = read_excel_data(MASTER_FILE)
    if master_data is None:
        logging.error("Masterファイルが読み込めませんでした")
        return False

    # dataディレクトリ内のExcelファイルを取得
    data_files = [f for f in os.listdir(DATA_DIR) if f.endswith(('.xlsx', '.xls'))]

    if not data_files:
        logging.info("処理対象のExcelファイルがありません")
        return True

    logging.info(f"処理対象ファイル数: {len(data_files)}")

    total_added = 0
    processed_files = 0

    for data_file in data_files:
        data_file_path = os.path.join(DATA_DIR, data_file)
        logging.info(f"--- {data_file} の処理開始 ---")

        # データファイルの読み込み
        new_data, new_headers = read_excel_data(data_file_path)
        if new_data is None:
            continue

        # データが空の場合はスキップ
        if not new_data:
            logging.warning(f"データが空です: {data_file}")
            continue

        # IDカラムを確認（設定値またはmasterの最初のカラムを使用）
        id_column = ID_COLUMN if ID_COLUMN in (master_headers or []) else (master_headers[0] if master_headers else new_headers[0])
        logging.info(f"IDカラム: {id_column}")

        # 新しいIDを取得
        new_ids = get_new_ids(master_data, new_data, id_column)

        if new_ids:
            # masterファイルを更新
            if update_master_excel(MASTER_FILE, new_data, new_ids, id_column, master_headers):
                total_added += len(new_ids)
                # masterデータを更新して次の処理に備える
                master_data, master_headers = read_excel_data(MASTER_FILE)
            else:
                logging.error(f"Masterファイルの更新に失敗: {data_file}")
                continue

        # 処理済みフォルダに移動
        if move_to_processed(data_file_path, PROCESSED_DIR):
            processed_files += 1

        logging.info(f"--- {data_file} の処理完了 ---")

    logging.info(f"=== 処理完了 ===")
    logging.info(f"処理ファイル数: {processed_files}")
    logging.info(f"追加レコード数: {total_added}")

    return True

def main():
    # 必要なディレクトリを作成
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    os.makedirs(LOG_DIR, exist_ok=True)

    # ファイルの存在確認
    if not os.path.exists(MASTER_FILE):
        print(f"Masterファイルが見つかりません: {MASTER_FILE}")
        return False

    if not os.path.exists(DATA_DIR):
        print(f"データディレクトリが見つかりません: {DATA_DIR}")
        return False

    # データ処理実行
    success = process_excel_files()

    if success:
        print("処理が正常に完了しました")
    else:
        print("処理中にエラーが発生しました")

    return success

if __name__ == "__main__":
    main()